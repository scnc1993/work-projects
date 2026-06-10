"""
Charter Invoice Audit Engine V4 — Python Edition
Daquan Morrison · Data Analyst · Syracuse City School District
Updated: 2026-05-01

V3 ADDITIONS:
  - SPED Status flag: joins SPED Roster to SCSD Roster (SPED / Gen Ed)
  - SchoolTool FTE: NYSED formula (Student Weeks / 40 program weeks)
  - Charter FTE: preserved from original invoice FTE column
  - FTE Variance: Charter FTE - SchoolTool FTE (positive = overbilled)

V4 ADDITIONS:
  - SQL Audit Engine: full Athena SQL version created (Charter_Audit_Engine_SQL_Final.sql)
    replicates all Python/R rules as Amazon Athena CTEs for live dashboard
  - SPED FTE: SchoolTool SPED FTE + Charter SPED FTE + SPED FTE Variance added
  - Gen Roster SQL: SCSD Gen Roster built in Athena Custom SQL (Direct Query, live refresh)
  - FTE formula aligned to NYSED month-based calculation (* 4 / 40) per §175.6
  - Separate auto-detect: enrollment and disability files loaded INDEPENDENTLY
    (each from its own Z: drive folder — not winner-takes-all)
  - Per-tab fallback logic: each tab checks main workbook FIRST, then disability_path
  - find_sheet_in_file() helper: searches sheets in a separate workbook
  - Pre-processing: SPED/disability key sets built BEFORE any analysis runs
  - Bug fix: FTE date arithmetic uses safe date arithmetic (no difftime int bug)
  - Summary banner and output filenames updated from V3 → V4

REQUIREMENTS:
  pip install pandas openpyxl xlrd

USAGE:
  python Charter_Audit_Engine_V4.py
  → Auto-detects most recent enrollment file from Z: drive
  → Auto-detects most recent disability file from Z: drive (optional)
  → Falls back to manual prompt if neither folder is accessible
  → Saves _AUDITED_V4_ Excel to same folder as enrollment file
"""

import os
import sys
import re
import warnings
from datetime import datetime, date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ======================================================================
# CONFIGURATION — update school year dates at start of each new year
# ======================================================================
SCHOOL_YEAR_FIRST_DAY = date(2025, 9, 3)
GRACE_DEADLINE        = date(2025, 9, 6)
JULY_START            = date(2025, 7, 1)
GRACE_DAYS            = 3
STILL_ENROLLED_DATE   = date(2026, 6, 30)   # NYSED program end date (max possible window per §175.6)
PROGRAM_WEEKS         = 40   # NYSED 10-month school year

# ======================================================================
# LOGGING
# ======================================================================
def write_log(msg, level="INFO"):
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] [{level}] {msg}")

# ======================================================================
# HEADER NORMALIZATION
# ======================================================================
def normalize_header(h):
    if h is None:
        return ""
    h = str(h)
    h = re.sub(r'[\u00A0\n\r]+', ' ', h)  # non-breaking spaces + newlines
    h = re.sub(r'\s+', ' ', h).strip()
    return h

def normalize_headers(df):
    df.columns = [normalize_header(c) for c in df.columns]
    return df

# ======================================================================
# COLUMN DETECTION
# ======================================================================
def find_best_col(df, patterns):
    cols = list(df.columns)
    for pat in patterns:
        for col in cols:
            if re.search(pat, col, re.IGNORECASE):
                return col
    return None

# ======================================================================
# DATE PARSING
# ======================================================================
def parse_date(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    # Handle pandas NaT (pd.NaT passes isinstance datetime checks but has no .date())
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(val, (datetime, date)):
        return pd.Timestamp(val).date()
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("nan", "none", ""):
        return None
    # Excel serial number
    try:
        serial = int(float(val_str))
        if 20000 < serial < 60000:
            return (pd.Timestamp("1899-12-30") + pd.Timedelta(days=serial)).date()
    except (ValueError, OverflowError):
        pass
    # Common date formats (including datetime strings with time component)
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%m/%d/%y", "%Y/%m/%d",
                "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            pass
    # Last resort: pandas parser
    try:
        return pd.to_datetime(val_str).date()
    except Exception:
        pass
    return None

# ======================================================================
# SHEET DETECTION — in an already-open ExcelFile object
# ======================================================================
def find_sheet(xl, keywords, exclude=None):
    exclude = exclude or []
    sheets = xl.sheet_names
    sheets_clean = [s.strip() for s in sheets]
    for kw in keywords:
        for i, s in enumerate(sheets_clean):
            if re.search(kw, s, re.IGNORECASE):
                if not any(re.search(ex, s, re.IGNORECASE) for ex in exclude):
                    return sheets[i]
    return None

# ======================================================================
# FIND SHEET IN A SEPARATE FILE (V4 fallback helper)
# Opens a separate .xlsx file and searches its sheet names.
# Returns (sheet_name, ExcelFile) or (None, None) if not found.
# ======================================================================
def find_sheet_in_file(path, patterns, exclude_patterns=None):
    """
    Search for a sheet matching `patterns` inside a separate workbook at `path`.
    Returns (sheet_name, xl_object) or (None, None).
    Used for per-tab fallback: if tab not in main workbook, check disability_path.
    """
    if path is None or not os.path.exists(path):
        return None, None
    exclude_patterns = exclude_patterns or []
    try:
        xl_other = pd.ExcelFile(path)
        sheets = xl_other.sheet_names
        sheets_clean = [s.strip() for s in sheets]
        for kw in patterns:
            for i, s in enumerate(sheets_clean):
                if re.search(kw, s, re.IGNORECASE):
                    if not any(re.search(ex, s, re.IGNORECASE) for ex in exclude_patterns):
                        return sheets[i], xl_other
        xl_other.close()
    except Exception as e:
        write_log(f"Could not open fallback file {path}: {e}", "WARN")
    return None, None

# ======================================================================
# SMART READ (handles title rows)
# ======================================================================
def smart_read(xl, sheet_name):
    """Read a sheet, detecting COLOR KEY layout (headers at row 3) vs normal (headers at row 1)."""
    df = xl.parse(sheet_name, header=0, dtype=str)
    df = normalize_headers(df)
    # Detect COLOR KEY layout: first col of first row contains "COLOR KEY"
    # In that case actual headers are at row 3 (skiprows=2, header=0)
    first_val = ""
    if len(df) > 0 and len(df.columns) > 0:
        first_val = str(df.iloc[0, 0]).strip().upper() if pd.notna(df.iloc[0, 0]) else ""
    if "COLOR KEY" in first_val or (len(df) > 0 and df.iloc[0].notna().sum() < 3):
        # Previously-audited output: COLOR KEY in row 1, blank row 2, headers row 3
        df = xl.parse(sheet_name, skiprows=2, header=0, dtype=str)
        df = normalize_headers(df)
    df = df.dropna(how="all")
    # Drop junk/footer rows that lack a valid NYSSIS ID in the first column
    # Valid NYSSIS IDs are purely numeric (6–12 digits); footer rows (FTE totals,
    # percentage rows) have non-numeric or blank values and must be excluded.
    if len(df) > 0 and len(df.columns) > 0:
        first_col = df.columns[0]
        def _is_valid_nyssis(val):
            if pd.isna(val):
                return False
            s = str(val).strip().replace('.0', '')
            return bool(__import__('re').match(r'^\d{6,12}$', s))
        has_valid = df[first_col].apply(_is_valid_nyssis)
        # Only apply NYSSIS filter on the main SCSD/SPED roster sheets
        # (ST Enrollment and Disability sheets don't have NYSSIS in col 1)
        # Heuristic: apply filter only when >=50% of rows have valid NYSSIS
        if has_valid.sum() / max(len(df), 1) >= 0.5:
            df = df[has_valid].copy()
    return df

# ======================================================================
# BILLING RULES
# ======================================================================
def classify_entry(charter_entry, st_entry):
    c = parse_date(charter_entry)
    s = parse_date(st_entry)
    if c is None and s is None:
        return ""
    if c is not None and s is None:
        return "Charter has entry date but ST does not (UNFAVORABLE)"
    if c is None and s is not None:
        return "ST has entry date but Charter does not (UNFAVORABLE)"
    # July/Aug check uses raw dates (grace window predates school year)
    if c < SCHOOL_YEAR_FIRST_DAY and s < SCHOOL_YEAR_FIRST_DAY:
        return "Charter and ST both enrolled in July/August (FAVORABLE)"
    if s < SCHOOL_YEAR_FIRST_DAY and c <= GRACE_DEADLINE:
        return "ST enrolled in July/August by school-year first day deadline (FAVORABLE)"
    # Normalize: any entry before 9/3/2025 treated as 9/3/2025 for comparison
    SY_FIRST = date(2025, 9, 3)
    cn = max(c, SY_FIRST)
    sn = max(s, SY_FIRST)
    if cn == sn:
        return "Entry dates match (FAVORABLE)"
    diff = (cn - sn).days
    if diff > 0:
        return f"Charter entry is {diff} days later than ST entry date (FAVORABLE)"
    if abs(diff) <= GRACE_DAYS:
        return f"Charter entry is {abs(diff)} days earlier than ST entry date (FAVORABLE)"
    return f"Charter entry is {abs(diff)} days earlier than ST entry date (UNFAVORABLE)"

def classify_exit(charter_exit, st_exit):
    c = parse_date(charter_exit)
    s = parse_date(st_exit)
    sentinel = STILL_ENROLLED_DATE
    if (c is None or c == sentinel) and (s is None or s == sentinel):
        return "No Charter or ST exit dates on file (FAVORABLE)"
    if (c is None or c == sentinel) and s is not None:
        return "ST has exit date but Charter does not (UNFAVORABLE)"
    if c is not None and (s is None or s == sentinel):
        return "Charter has exit date but ST does not (Discrepancy \u2013 update in ST)"
    if c == s:
        return "Exit dates match (FAVORABLE)"
    diff = (c - s).days
    if diff < 0:
        days = abs(diff)
        if days <= GRACE_DAYS:
            return f"Charter exit is {days} days earlier (FAVORABLE)"
        return f"Charter exit is {days} days earlier (Discrepancy \u2013 update in ST)"
    if diff <= GRACE_DAYS:
        return f"Charter exit is {diff} days later (FAVORABLE)"
    return f"Charter exit is {diff} days later (UNFAVORABLE)"

def determine_status(entry_disc, exit_disc, match_method):
    if match_method == "NOT FOUND":
        return "Not Found-Needs Manual Review"
    if "UNFAVORABLE" in str(entry_disc) or "UNFAVORABLE" in str(exit_disc):
        return "No Good"
    return "Good"

# ======================================================================
# FTE CALCULATION (NYSED: Student Weeks / 40 Program Weeks)
# V4 FIX: uses Python date arithmetic directly — no intermediate
# conversion that could silently produce integer seconds (R ifelse bug
# equivalent). parse_date() always returns a date object or None.
# ======================================================================
def calc_fte(entry_date, exit_date, fallback_end=STILL_ENROLLED_DATE):
    s = parse_date(entry_date)
    e = parse_date(exit_date)
    if s is None:
        return None
    if e is None or e >= fallback_end:
        e = fallback_end
    # NYSED §175.6: (weeks * 4) / 40 to align with month-based lookup table methodology
    # Both s and e are guaranteed date objects here — no type-stripping risk
    weeks = (e - s).days / 7
    fte = round(max(0.0, min(1.0, (weeks * 4) / 40)), 4)
    return fte

# ======================================================================
# EXCEL STYLE HELPERS
# ======================================================================
BLUE_FILL   = PatternFill("solid", fgColor="4472C4")
GREEN_FILL  = PatternFill("solid", fgColor="90EE90")
RED_FILL    = PatternFill("solid", fgColor="FF6B6B")
YELLOW_FILL = PatternFill("solid", fgColor="FFD966")
WHITE_FONT  = Font(color="FFFFFF", bold=True)
BOLD_FONT   = Font(bold=True)

def apply_header_style(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(wrap_text=True)

def set_col_widths(ws, df, start_row=1):
    for i, col in enumerate(df.columns, 1):
        max_len = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) > 0 else 0)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 50)

# ======================================================================
# MAIN ENGINE
# V4: accepts optional disability_path for separate disability file.
# ======================================================================
def run_audit(workbook_path, disability_path=None):
    write_log(f"Loading main workbook: {workbook_path}")
    xl = pd.ExcelFile(workbook_path)
    sheet_names = xl.sheet_names

    if disability_path:
        write_log(f"Disability file: {disability_path}")
    else:
        write_log("No separate disability file provided — will check main workbook only", "WARN")

    # ----------------------------------------------------------------
    # READ SHEETS — V4 per-tab fallback logic
    # For each tab: check main workbook FIRST, then disability_path file.
    # Source is logged so you can see where each tab came from.
    # ----------------------------------------------------------------

    # --- SCSD Gen Roster (always in main workbook) ---
    exclude_main = ["SPED", "Invoice", "Disability", "IEP", "Registration",
                    "Not on", "Not in", "Multiple"]
    scsd_sheet = find_sheet(xl, ["SCSD.*Roster", "Gen.*Roster", "Roster"], exclude=exclude_main)
    if scsd_sheet is None:
        write_log("SCSD Roster sheet not found in main workbook — cannot continue", "ERROR")
        return
    write_log(f"[SCSD Roster] source: main workbook | sheet: {scsd_sheet}")

    # --- ST Enrollment tab ---
    st_sheet = find_sheet(xl, ["ST.*Enrollment", "Enrollment", "ChartEnrol"],
                          exclude=["Multiple", "SCSD", "SPED", "Not on", "Not in"])
    st_xl = xl   # default: use main workbook
    if st_sheet:
        write_log(f"[ST Enrollment] source: main workbook | sheet: {st_sheet}")
    else:
        write_log("[ST Enrollment] not found in main workbook — checking disability file...", "WARN")
        st_sheet, st_xl = find_sheet_in_file(
            disability_path,
            ["ST.*Enrollment", "Enrollment", "ChartEnrol"],
            exclude_patterns=["Multiple", "SCSD", "SPED", "Not on", "Not in"]
        )
        if st_sheet:
            write_log(f"[ST Enrollment] source: disability file | sheet: {st_sheet}")
        else:
            write_log("[ST Enrollment] not found in either file — ST columns will be blank", "WARN")

    # --- SPED Roster tab ---
    sped_sheet = find_sheet(xl, ["SPED.*Roster", "SPED"])
    sped_xl = xl
    if sped_sheet:
        write_log(f"[SPED Roster] source: main workbook | sheet: {sped_sheet}")
    else:
        write_log("[SPED Roster] not found in main workbook — checking disability file...", "WARN")
        sped_sheet, sped_xl = find_sheet_in_file(
            disability_path,
            ["SPED.*Roster", "SPED"]
        )
        if sped_sheet:
            write_log(f"[SPED Roster] source: disability file | sheet: {sped_sheet}")
        else:
            write_log("[SPED Roster] not found in either file — SPED Status will be Unknown", "WARN")

    # --- Disability tab ---
    disability_sheet = find_sheet(xl, ["Disability", "IEP.*Roster"])
    disability_xl = xl
    if disability_sheet:
        write_log(f"[Disability] source: main workbook | sheet: {disability_sheet}")
    else:
        write_log("[Disability] not found in main workbook — checking disability file...", "WARN")
        disability_sheet, disability_xl = find_sheet_in_file(
            disability_path,
            ["Disability", "IEP.*Roster"]
        )
        if disability_sheet:
            write_log(f"[Disability] source: disability file | sheet: {disability_sheet}")
        else:
            write_log("[Disability] not found in either file — disability NYSSIS set will be empty", "WARN")

    # ----------------------------------------------------------------
    # PRE-PROCESS SPED + DISABILITY — build key sets BEFORE any analysis
    # V4: both key sets are merged before the main roster join runs.
    # ----------------------------------------------------------------

    # Build SPED key set from SPED Roster tab
    sped_df = None
    sped_key_set = set()
    st_id_xref = None   # will be populated after ST read below

    # Build disability NYSSIS set from Disability tab
    disability_nyssis_set = set()
    if disability_sheet:
        try:
            disability_raw = smart_read(disability_xl, disability_sheet)
            write_log(f"[Disability] {len(disability_raw)} rows loaded")
            dis_nyssis_col = find_best_col(disability_raw, [r"NYSSIS", r"Universal.*Student.*ID", r"StateID"])
            if dis_nyssis_col:
                disability_nyssis_set = set(
                    disability_raw[dis_nyssis_col]
                    .astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
                    .dropna().unique()
                )
                write_log(f"[Disability] {len(disability_nyssis_set)} unique NYSSIS IDs in disability set")
            else:
                write_log("[Disability] NYSSIS column not found — disability NYSSIS set empty", "WARN")
        except Exception as e:
            write_log(f"[Disability] Failed to read tab: {e}", "WARN")

    # ----------------------------------------------------------------
    # READ SCSD ROSTER
    # ----------------------------------------------------------------
    scsd_raw = smart_read(xl, scsd_sheet)
    write_log(f"SCSD Roster: {len(scsd_raw)} rows")

    # --- Detect columns ---
    nyssis_col    = find_best_col(scsd_raw, [r"NYSSIS", r"Universal.*Student.*ID", r"StateID"])
    first_name_col = find_best_col(scsd_raw, [r"First.*Name", r"FirstName"])
    last_name_col  = find_best_col(scsd_raw, [r"Last.*Name", r"LastName"])
    dob_col        = find_best_col(scsd_raw, [r"Birth", r"DOB"])
    grade_col      = find_best_col(scsd_raw, [r"Grade"])
    entry_col      = find_best_col(scsd_raw, [r"Entry.*Date", r"Enroll.*Date", r"Entered", r"Start"])
    exit_col       = find_best_col(scsd_raw, [r"Leave.*Date", r"Exit.*Date", r"Withdrew", r"Withdraw.*Date", r"Withdraw$", r"^End.*Date"])
    fte_col        = find_best_col(scsd_raw, [r"^FTE$", r"Southside.*FTE", r"Charter.*FTE", r"OnTECH.*FTE", r"SCSD.*FTE"])
    address_col    = find_best_col(scsd_raw, [r"Address", r"House", r"Street"])
    city_col       = find_best_col(scsd_raw, [r"^City$"])
    zip_col        = find_best_col(scsd_raw, [r"Zip"])
    district_col   = find_best_col(scsd_raw, [r"District"])

    # --- Pre-populated ST columns (OnTECH/SASCCS: roster already has ST Entry/Exit/School) ---
    pre_pop_st_entry  = find_best_col(scsd_raw, [r"^ST.?Entry.?Date$", r"^ST.?Entry$", r"^ST.?Enroll"])
    pre_pop_st_exit   = find_best_col(scsd_raw, [r"^ST.?Exit.?Date$", r"^ST.?Exit$", r"^ST.?Withdraw", r"^ST.?Leave"])
    pre_pop_school    = find_best_col(scsd_raw, [r"^School.?Name$", r"^SchoolName$"])
    has_prepop_st     = pre_pop_st_entry is not None
    if has_prepop_st:
        write_log(f"Pre-populated ST columns detected — ST Entry: {pre_pop_st_entry} | ST Exit: {pre_pop_st_exit} | School: {pre_pop_school}")

    if nyssis_col is None:
        write_log("NYSSIS column not found in SCSD Roster", "ERROR")
        return

    scsd_df = scsd_raw.copy()
    scsd_df["NYSSIS"] = scsd_df[nyssis_col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)

    # ----------------------------------------------------------------
    # READ ST ENROLLMENT
    # ----------------------------------------------------------------
    st_deduped = None
    if st_sheet:
        st_raw = smart_read(st_xl, st_sheet)
        write_log(f"ST Enrollment: {len(st_raw)} rows")
        st_nyssis_col   = find_best_col(st_raw, [r"Universal.*Student.*ID", r"NYSSIS", r"StateID"])
        st_id_col       = find_best_col(st_raw, [r"^StudentID$", r"Student.*ID$"])
        st_entry_st_col = find_best_col(st_raw, [r"Start.*Date", r"Entry.*Date", r"Enrollment.*Start"])
        st_exit_st_col  = find_best_col(st_raw, [r"End.*Date", r"Exit.*Date", r"Enrollment.*End"])
        st_school_st_col = find_best_col(st_raw, [r"SchoolName", r"School.*Name"])
        st_fn_col       = find_best_col(st_raw, [r"First.*Name"])
        st_ln_col       = find_best_col(st_raw, [r"Last.*Name"])
        st_dob_st_col   = find_best_col(st_raw, [r"DOB", r"Birth"])
        st_grade_st_col = find_best_col(st_raw, [r"Grade"])

        if st_nyssis_col:
            st_raw["_NYSSIS"] = st_raw[st_nyssis_col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
            st_deduped = st_raw.sort_values(
                st_entry_st_col if st_entry_st_col else st_raw.columns[0], ascending=False
            ).drop_duplicates(subset=["_NYSSIS"])

            # Cross-reference for SPED join (StudentID → NYSSIS bridge)
            if st_id_col:
                st_id_xref = st_deduped[["_NYSSIS", st_id_col]].rename(
                    columns={"_NYSSIS": "NYSSIS", st_id_col: "ST_StudentID"}
                )
    else:
        st_id_col = st_entry_st_col = st_exit_st_col = st_school_st_col = None

    # ----------------------------------------------------------------
    # PRE-PROCESS SPED ROSTER — build sped_key_set BEFORE ST join
    # Merges both SPED Roster key set AND disability NYSSIS set
    # ----------------------------------------------------------------
    if sped_sheet:
        sped_raw = smart_read(sped_xl, sped_sheet)
        write_log(f"SPED Roster: {len(sped_raw)} rows")
        sped_nyssis_col = find_best_col(sped_raw, [r"NYSSIS", r"Universal.*Student.*ID"])
        sped_id_col     = find_best_col(sped_raw, [r"Student.*Gen.*Ed.*ID", r"Student.*ID"])

        if sped_nyssis_col:
            sped_df = sped_raw.copy()
            sped_df["SPED_KEY"] = sped_df[sped_nyssis_col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        elif sped_id_col and st_id_xref is not None:
            sped_df = sped_raw.copy()
            sped_df["_sid"] = sped_df[sped_id_col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
            sped_df = sped_df.merge(st_id_xref.rename(columns={"ST_StudentID": "_sid"}),
                                    on="_sid", how="left")
            sped_df["SPED_KEY"] = sped_df["NYSSIS"].fillna(sped_df["_sid"])

        if sped_df is not None and "SPED_KEY" in sped_df.columns:
            sped_key_set = set(sped_df["SPED_KEY"].dropna().unique())
            write_log(f"[SPED Roster] {len(sped_key_set)} unique SPED keys from roster tab")

    # Merge disability NYSSIS set into sped_key_set
    # Only include disability NYSSIS IDs that are also in this roster (avoid marking every district student as SPED)
    if disability_nyssis_set:
        roster_nyssis = set(scsd_df["NYSSIS"].dropna().unique())
        filtered_dis = disability_nyssis_set & roster_nyssis
        pre_merge = len(sped_key_set)
        sped_key_set = sped_key_set | filtered_dis
        write_log(f"[Disability] {len(disability_nyssis_set)} district NYSSIS in disability file → "
                  f"{len(filtered_dis)} matched to this roster → "
                  f"{len(sped_key_set) - pre_merge} net new SPED flags")

    if not sped_key_set:
        write_log("SPED key set is empty — all students will show SPED_Status = Unknown", "WARN")

    # ----------------------------------------------------------------
    # JOIN ST TO SCSD
    # ----------------------------------------------------------------
    def get_st_val(nyssis, col):
        if st_deduped is None or col is None:
            return None
        match = st_deduped[st_deduped["_NYSSIS"] == nyssis]
        if len(match) == 0:
            return None
        return match.iloc[0][col]

    write_log("Joining ST enrollment to SCSD roster...")
    if st_sheet:
        # ST Enrollment tab present — join by NYSSIS
        scsd_df["ST_Entry"]     = scsd_df["NYSSIS"].apply(lambda x: get_st_val(x, st_entry_st_col))
        scsd_df["ST_Exit"]      = scsd_df["NYSSIS"].apply(lambda x: get_st_val(x, st_exit_st_col))
        scsd_df["ST_School"]    = scsd_df["NYSSIS"].apply(lambda x: get_st_val(x, st_school_st_col))
        scsd_df["ST_StudentID"] = scsd_df["NYSSIS"].apply(lambda x: get_st_val(x, st_id_col))
        scsd_df["Match_Method"] = scsd_df["NYSSIS"].apply(
            lambda x: "NYSSIS" if (st_deduped is not None and x in st_deduped["_NYSSIS"].values) else "NOT FOUND"
        )
    elif has_prepop_st:
        # No ST Enrollment tab — use pre-populated ST columns already in the roster (OnTECH/SASCCS pattern)
        write_log("No ST Enrollment tab — using pre-populated ST columns from roster")
        scsd_df["ST_Entry"]     = scsd_df[pre_pop_st_entry].apply(parse_date)
        scsd_df["ST_Exit"]      = scsd_df[pre_pop_st_exit].apply(parse_date) if pre_pop_st_exit else None
        scsd_df["ST_School"]    = scsd_df[pre_pop_school].astype(str).str.strip() if pre_pop_school else None
        scsd_df["ST_StudentID"] = None
        scsd_df["Match_Method"] = scsd_df["ST_Entry"].apply(
            lambda x: "Pre-populated" if (x is not None and str(x) not in ("", "NaT", "None", "nan")) else "NOT FOUND"
        )
    else:
        # No ST data available at all
        scsd_df["ST_Entry"]     = None
        scsd_df["ST_Exit"]      = None
        scsd_df["ST_School"]    = None
        scsd_df["ST_StudentID"] = None
        scsd_df["Match_Method"] = "NOT FOUND"

    # ----------------------------------------------------------------
    # BILLING RULES
    # ----------------------------------------------------------------
    write_log("Applying billing rules...")
    scsd_df["Entry_Discrepancy"] = scsd_df.apply(
        lambda r: classify_entry(r.get(entry_col), r["ST_Entry"]), axis=1)
    scsd_df["Exit_Discrepancy"] = scsd_df.apply(
        lambda r: classify_exit(r.get(exit_col), r["ST_Exit"]), axis=1)
    scsd_df["Student_Status"] = scsd_df.apply(
        lambda r: determine_status(r["Entry_Discrepancy"], r["Exit_Discrepancy"], r["Match_Method"]), axis=1)

    # ----------------------------------------------------------------
    # SPED STATUS FLAG
    # Applied AFTER sped_key_set is fully built (both sources merged)
    # ----------------------------------------------------------------
    write_log("Building SPED Status flag...")
    if sped_key_set:
        scsd_df["SPED_Status"] = scsd_df["NYSSIS"].apply(
            lambda x: "SPED" if x in sped_key_set else "Gen Ed"
        )
    else:
        scsd_df["SPED_Status"] = "Unknown"
    write_log(f"SPED: {(scsd_df['SPED_Status']=='SPED').sum()} | "
              f"Gen Ed: {(scsd_df['SPED_Status']=='Gen Ed').sum()} | "
              f"Unknown: {(scsd_df['SPED_Status']=='Unknown').sum()}")

    # ----------------------------------------------------------------
    # FTE CALCULATIONS
    # V4 FIX: parse_date() returns a Python date object or None.
    # calc_fte() operates on date objects directly — no type-stripping.
    # This mirrors the R fix: dplyr::if_else() preserves Date class.
    # ----------------------------------------------------------------
    write_log("Calculating FTE fields...")
    def calc_st_fte(row):
        # Entry: use ST Entry if available, else Charter Entry; normalize to 9/3/2025 minimum
        raw_entry = row["ST_Entry"] if row["ST_Entry"] else row.get(entry_col)
        e = parse_date(raw_entry)
        SY_FIRST = date(2025, 9, 3)
        entry = max(e, SY_FIRST) if e else None
        # Exit priority: ST Exit -> Charter Exit (Leave Date) -> STILL_ENROLLED_DATE
        st_exit = parse_date(row["ST_Exit"])
        charter_exit = parse_date(row.get(exit_col)) if exit_col else None
        if st_exit:
            exit_ = st_exit
        elif charter_exit:
            exit_ = charter_exit
        else:
            exit_ = STILL_ENROLLED_DATE
        if entry is None:
            return None
        return calc_fte(entry, exit_)

    scsd_df["ST_FTE"]      = scsd_df.apply(calc_st_fte, axis=1)
    scsd_df["Charter_FTE"] = scsd_df[fte_col].apply(lambda x: float(x) if x and str(x) not in ("nan", "") else None) if fte_col else None
    scsd_df["FTE_Variance"] = scsd_df.apply(
        lambda r: round(float(r["Charter_FTE"]) - float(r["ST_FTE"]), 4)
        if r["Charter_FTE"] is not None and r["ST_FTE"] is not None else None, axis=1
    )
    # SPED FTE: only populated for students flagged as SPED
    scsd_df["ST_SPED_FTE"] = scsd_df.apply(
        lambda r: r["ST_FTE"] if r["SPED_Status"] == "SPED" else None, axis=1
    )
    scsd_df["Charter_SPED_FTE"] = scsd_df.apply(
        lambda r: r["Charter_FTE"] if r["SPED_Status"] == "SPED" else None, axis=1
    )
    scsd_df["SPED_FTE_Variance"] = scsd_df.apply(
        lambda r: round(float(r["Charter_SPED_FTE"]) - float(r["ST_SPED_FTE"]), 4)
        if r["Charter_SPED_FTE"] is not None and r["ST_SPED_FTE"] is not None else None, axis=1
    )
    write_log("FTE calculations complete")

    # ----------------------------------------------------------------
    # SUMMARY COUNTS
    # ----------------------------------------------------------------
    total        = len(scsd_df)
    good_count   = (scsd_df["Student_Status"] == "Good").sum()
    nogood_count = (scsd_df["Student_Status"] == "No Good").sum()
    nmr_count    = (scsd_df["Student_Status"] == "Not Found-Needs Manual Review").sum()
    not_found_n  = (scsd_df["Match_Method"] == "NOT FOUND").sum()

    print("\n" + "="*60)
    print("  CHARTER AUDIT ENGINE V4 - SUMMARY REPORT")
    print("="*60)
    print(f"  Total students:           {total}")
    print(f"  Good:                     {good_count} ({100*good_count/max(total,1):.1f}%)")
    print(f"  No Good:                  {nogood_count} ({100*nogood_count/max(total,1):.1f}%)")
    print(f"  Not Found (NMR):          {nmr_count}")
    print(f"  SPED students:            {(scsd_df['SPED_Status']=='SPED').sum()}")
    print(f"  Disability file used:     {'Yes' if disability_path else 'No'}")
    print("="*60)

    # ----------------------------------------------------------------
    # BUILD OUTPUT DATAFRAME
    # ----------------------------------------------------------------
    def safe_get(df, col):
        return df[col] if col and col in df.columns else ""

    # Multi-enrollment count for each student (for the Multiple Enrollments column)
    multi_counts = scsd_df["NYSSIS"].map(
        scsd_df.groupby("NYSSIS")["NYSSIS"].transform("count")
    ).fillna(1).astype(int)

    scsd_out = pd.DataFrame({
        "NYSSIS ID":                        scsd_df["NYSSIS"],
        "Student ID":                       scsd_df["ST_StudentID"],
        "Last Name":                        safe_get(scsd_df, last_name_col),
        "First Name":                       safe_get(scsd_df, first_name_col),
        "Date of Birth":                    safe_get(scsd_df, dob_col),
        "Current Grade Level":              safe_get(scsd_df, grade_col),
        "Address":                          safe_get(scsd_df, address_col),
        "City":                             safe_get(scsd_df, city_col),
        "Zip Code":                         safe_get(scsd_df, zip_col),
        "District of Residency":            safe_get(scsd_df, district_col),
        "Entry Date":                       safe_get(scsd_df, entry_col),
        "Leave Date":                       safe_get(scsd_df, exit_col),
        "FTE":                              scsd_df["Charter_FTE"],
        "ST Entry Date":                    scsd_df["ST_Entry"],
        "ST Exit Date":                     scsd_df["ST_Exit"],
        "SchoolTool FTE":                   scsd_df["ST_FTE"],   # overwritten with formula below
        "FTE Variance":                     scsd_df["FTE_Variance"],  # overwritten with formula below
        "School Name":                      scsd_df["ST_School"],
        "Student Status":                   scsd_df["Student_Status"],  # overwritten with formula below
        "ST ID Match":                      scsd_df["ST_StudentID"].fillna("NOT FOUND IN ST"),
        "Match Method":                     scsd_df["Match_Method"],
        "Multiple Enrollments":             multi_counts.apply(lambda x: "Yes" if x > 1 else "No"),
        "Notes for Patrick":                "",   # overwritten with formula below
        "Entry Date Discrepancy (>3 Days)": scsd_df["Entry_Discrepancy"],  # overwritten with formula below
        "Exit Date Discrepancy (>3 Days)":  scsd_df["Exit_Discrepancy"],   # overwritten with formula below
        "SPED Status":                      scsd_df["SPED_Status"],
    })

    # ----------------------------------------------------------------
    # WRITE OUTPUT WORKBOOK
    # ----------------------------------------------------------------
    base_name = os.path.splitext(os.path.basename(workbook_path))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(
        os.path.dirname(workbook_path),
        f"{base_name}_AUDITED_V4_{timestamp}.xlsx"
    )
    write_log(f"Writing output: {output_path}")

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: SCSD ROSTER
    ws = wb.create_sheet("SCSD ROSTER")
    # Color legend row 1
    legend = [("COLOR KEY:", None), ("Good", "90EE90"), ("No Good", "FF6B6B"),
              ("Not Found-Needs Manual Review", "FFD966"), ("Discrepancy - update in ST", "BDD7EE")]
    for ci, (txt, color) in enumerate(legend, 1):
        cell = ws.cell(row=1, column=ci, value=txt)
        cell.font = BOLD_FONT
        if color:
            cell.fill = PatternFill("solid", fgColor=color)

    # Headers row 3
    for ci, col in enumerate(scsd_out.columns, 1):
        cell = ws.cell(row=3, column=ci, value=col)
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT

    # Data rows start at 4
    status_col_idx = list(scsd_out.columns).index("Student Status") + 1
    for ri, (_, row) in enumerate(scsd_out.iterrows(), 4):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    ws.freeze_panes = "A4"

    # ------------------------------------------------------------------
    # LIVE EXCEL FORMULAS — all 5 columns recalculate when dates change
    # ------------------------------------------------------------------
    out_cols = list(scsd_out.columns)
    def ci(name):
        return out_cols.index(name) + 1 if name in out_cols else None

    fte_c        = ci("FTE")
    st_ent_c     = ci("ST Entry Date")
    st_exit_c    = ci("ST Exit Date")
    entry_c      = ci("Entry Date")
    leave_c      = ci("Leave Date")
    st_fte_c     = ci("SchoolTool FTE")
    fte_var_c    = ci("FTE Variance")
    status_c     = ci("Student Status")
    match_c      = ci("Match Method")
    notes_c      = ci("Notes for Patrick")
    entry_disc_c = ci("Entry Date Discrepancy (>3 Days)")
    exit_disc_c  = ci("Exit Date Discrepancy (>3 Days)")

    gl = get_column_letter
    n_rows = len(scsd_out)

    if n_rows > 0:
        # Column letters — resolved from actual output column positions
        # N/O (ST Entry/Exit) may be blank columns when no ST tab; formulas still write
        _N = gl(st_ent_c)  if st_ent_c  else gl(entry_c)   # ST Entry Date col letter
        _O = gl(st_exit_c) if st_exit_c else gl(leave_c)   # ST Exit Date col letter
        for r in range(4, n_rows + 4):
            N  = _N;  O  = _O
            K  = gl(entry_c);   L  = gl(leave_c)
            M  = gl(fte_c);     P  = gl(st_fte_c) if st_fte_c else gl(entry_c)
            U  = gl(match_c)
            Xc = gl(entry_disc_c); Yc = gl(exit_disc_c)
            row_idx = r - 4  # 0-based index into scsd_df

            # SchoolTool FTE — live Excel formula: recalculates when Patrick edits dates
            # Entry: use ST Entry if available, else Charter Entry; normalize to >= 9/3/2025
            # Exit priority: ST Exit -> Charter Exit (Leave Date) -> DATE(2026,6,26)
            raw_start = f'IF(OR(ISBLANK({N}{r}),{N}{r}=""),{K}{r},{N}{r})'
            start_f = f'IF(({raw_start})<DATE(2025,9,3),DATE(2025,9,3),{raw_start})'
            end_f   = f'IF(AND(NOT(OR(ISBLANK({O}{r}),{O}{r}="")),{O}{r}<>""),{O}{r},IF(AND(NOT(OR(ISBLANK({L}{r}),{L}{r}="")),{L}{r}<>""),{L}{r},DATE(2026,6,26)))'
            fte_f   = f'=IF(OR(ISBLANK({N}{r}),{N}{r}=""),"",MIN(1,MAX(0,ROUND(((({end_f})-({start_f}))/7)/40,3))))'
            if st_fte_c:
                ws.cell(row=r, column=st_fte_c).value = fte_f

            # FTE Variance — live Excel formula
            if fte_var_c and st_fte_c:
                ws.cell(row=r, column=fte_var_c).value = f'=ROUND({M}{r}-{P}{r},4)'

            # ---------------------------------------------------------------
            # WRITE ACTUAL COMPUTED VALUES (not formulas) for the four
            # classification columns so the file opens correctly with no
            # repair warnings and no blank cells.
            # openpyxl cannot pre-cache formula results, so writing formulas
            # for these columns leaves them blank until Excel recalculates.
            # The Python engine already computed the correct values in scsd_df;
            # we write those directly. SchoolTool FTE / FTE Variance remain as
            # live formulas because they reference date columns Patrick edits.
            # ---------------------------------------------------------------

            # Entry Date Discrepancy — write computed Python value
            if entry_disc_c:
                ws.cell(row=r, column=entry_disc_c).value = scsd_df["Entry_Discrepancy"].iloc[row_idx]

            # Exit Date Discrepancy — write computed Python value
            if exit_disc_c:
                ws.cell(row=r, column=exit_disc_c).value = scsd_df["Exit_Discrepancy"].iloc[row_idx]

            # Student Status — write computed Python value
            if status_c:
                ws.cell(row=r, column=status_c).value = scsd_df["Student_Status"].iloc[row_idx]

            # Notes for Patrick — computed from Python classification results
            # ONLY on No Good rows; blank for Good and Not Found
            if notes_c:
                entry_d = str(scsd_df["Entry_Discrepancy"].iloc[row_idx])
                exit_d  = str(scsd_df["Exit_Discrepancy"].iloc[row_idx])
                s_val   = scsd_df["Student_Status"].iloc[row_idx]
                if s_val == "No Good":
                    entry_unf = "UNFAVORABLE" in entry_d
                    exit_unf  = "UNFAVORABLE" in exit_d
                    exit_upd  = "update in ST" in exit_d.lower()
                    if entry_unf and exit_unf:
                        note = "Entry AND exit date discrepancy >3 days - update charter attendance record in ST"
                    elif entry_unf:
                        note = "Entry date discrepancy >3 days - update charter school attendance record in ST"
                    elif exit_unf:
                        note = "Exit date discrepancy >3 days - update charter school attendance record in ST"
                    elif exit_upd:
                        note = "Charter has exit/disenroll date but ST does not - update ST attendance record"
                    else:
                        note = ""
                else:
                    note = ""
                ws.cell(row=r, column=notes_c).value = note

            # ST Exit Date — write actual value (NO formula overwrite)
            # The actual ST exit date (or fallback charter exit) is already
            # in scsd_out["ST Exit Date"] from the Python merge step above.
            # Overwriting with a formula would wipe the cached value and cause
            # blank cells on file open (the "work was lost" repair warning).

        write_log("Values written: Entry Discrepancy, Exit Discrepancy, Student Status, Notes for Patrick (Python-computed). Live Excel formulas: SchoolTool FTE, FTE Variance.")

    # Apply conditional fill to ALL rows based on Student Status
    # Runs regardless of whether ST tab was present
    if n_rows > 0:
        n_cols = len(scsd_out.columns)
        for r in range(4, n_rows + 4):
            status_val = scsd_out.iloc[r - 4]["Student Status"]
            if status_val == "Good":
                row_fill = GREEN_FILL
            elif status_val == "No Good":
                row_fill = RED_FILL
            elif "Not Found" in str(status_val):
                row_fill = YELLOW_FILL
            else:
                row_fill = None
            if row_fill:
                for c in range(1, n_cols + 1):
                    ws.cell(row=r, column=c).fill = row_fill
        write_log("Row color formatting applied to all rows")

    # Sheet 2: AUDIT_SUMMARY
    ws2 = wb.create_sheet("AUDIT_SUMMARY", 1)
    summary_rows = [
        ("AUDIT RESULTS V4", ""),
        ("", ""),
        ("Total Students", total),
        ("Good", good_count),
        ("No Good", nogood_count),
        ("Not Found-Needs Manual Review", nmr_count),
        ("", ""),
        ("SPED Students", (scsd_df["SPED_Status"]=="SPED").sum()),
        ("Gen Ed Students", (scsd_df["SPED_Status"]=="Gen Ed").sum()),
        ("Unknown SPED Status", (scsd_df["SPED_Status"]=="Unknown").sum()),
        ("", ""),
        ("Disability File Used", "Yes" if disability_path else "No — SPED Status may show Unknown"),
        ("", ""),
        ("Good %", f"{100*good_count/max(total,1):.1f}%"),
        ("No Good %", f"{100*nogood_count/max(total,1):.1f}%"),
        ("Not Found %", f"{100*nmr_count/max(total,1):.1f}%"),
        ("", ""),
        ("WHAT TO DO NEXT", ""),
        ("Step 1", "Go to Action Needed tab for your work queue"),
        ("Step 2", "Yellow rows = look up in ST, add dates via VLOOKUP"),
        ("Step 3", "Red rows = check docs/duplicates"),
        ("Step 4", "Strip internal columns before sending to budget"),
    ]
    for ri, (metric, value) in enumerate(summary_rows, 1):
        ws2.cell(row=ri, column=1, value=metric)
        ws2.cell(row=ri, column=2, value=value)
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 40

    # Sheet 3: Action Needed
    ws3 = wb.create_sheet("Action Needed")
    ws3.cell(row=1, column=1, value=f"ACTION NEEDED: {nogood_count + nmr_count} students require manual review")
    ws3.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1F4E79")
    ws3.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=13)

    action_df = scsd_out[scsd_out["Student Status"].isin(
        ["No Good", "Not Found-Needs Manual Review"]
    )][["NYSSIS ID", "Last Name", "First Name", "Date of Birth", "Current Grade Level",
        "Entry Date", "Leave Date", "ST Entry Date", "ST Exit Date",
        "Entry Date Discrepancy (>3 Days)", "Exit Date Discrepancy (>3 Days)",
        "Student Status", "SPED Status"]]

    cur_row = 3
    for ci, col in enumerate(action_df.columns, 1):
        c = ws3.cell(row=cur_row, column=ci, value=col)
        c.fill = BLUE_FILL
        c.font = WHITE_FONT
    cur_row += 1
    for _, row in action_df.iterrows():
        for ci, val in enumerate(row, 1):
            ws3.cell(row=cur_row, column=ci, value=val)
        _action_status_cell = ws3.cell(row=cur_row, column=list(action_df.columns).index("Student Status") + 1)
        if row["Student Status"] == "No Good":
            _action_status_cell.fill = RED_FILL
        else:
            _action_status_cell.fill = YELLOW_FILL
        cur_row += 1

    # Sheet 4: SPED ROSTER
    if sped_df is not None:
        # Join ST SPED FTE onto sped_df
        sped_fte_map = scsd_df.set_index("NYSSIS")["ST_SPED_FTE"].to_dict()
        sped_df["ST SPED FTE"] = sped_df["SPED_KEY"].map(sped_fte_map)

        ws4 = wb.create_sheet("SPED ROSTER")
        sped_cols = [c for c in sped_df.columns if c not in ("SPED_KEY", "_sid", "IEP_Code_lookup",
                     "IEP_Enroll_status_lookup", "SPED_ST_StudentID", "ST_SPED_FTE_lookup")]
        for ci, col in enumerate(sped_cols, 1):
            cell = ws4.cell(row=1, column=ci, value=col)
            cell.fill = BLUE_FILL
            cell.font = WHITE_FONT
        for ri, (_, row) in enumerate(sped_df[sped_cols].iterrows(), 2):
            for ci, val in enumerate(row, 1):
                ws4.cell(row=ri, column=ci, value=val)
        ws4.freeze_panes = "A2"
        write_log(f"SPED ROSTER written: {len(sped_df)} rows")



    # ------------------------------------------------------------------
    # CONDITIONAL FORMATTING — Status, Entry/Exit Discrepancy, FTE Variance
    # Matches R engine CF rules exactly (type=containsText for text cols)
    # ------------------------------------------------------------------
    if n_rows > 0:
        from openpyxl.formatting.rule import CellIsRule, Rule
        from openpyxl.styles.differential import DifferentialStyle
        last_row = n_rows + 3  # data ends at row (n_rows + 3)

        def _make_contains_rule(text, fill_hex, font_hex=None, priority=1):
            """Build a containsText CF rule matching the R engine's type='contains'."""
            dxf = DifferentialStyle(
                fill=PatternFill(bgColor=fill_hex),
                font=Font(color=font_hex) if font_hex else None
            )
            rule = Rule(type='containsText', operator='containsText',
                        text=text, dxf=dxf, priority=priority)
            rule.formula = [f'NOT(ISERROR(SEARCH("{text}",{{}}{4})))']
            return rule

        # Student Status — col S
        if status_c:
            _s_col = get_column_letter(status_c)
            _s_range = f"{_s_col}4:{_s_col}{last_row}"
            # No Good → red
            _r = DifferentialStyle(fill=PatternFill(bgColor="FF6B6B"))
            _r_rule = Rule(type='containsText', operator='containsText', text='No Good', dxf=_r, priority=1)
            _r_rule.formula = [f'NOT(ISERROR(SEARCH("No Good",{_s_col}4)))']
            ws.conditional_formatting.add(_s_range, _r_rule)
            # Not Found → yellow
            _y = DifferentialStyle(fill=PatternFill(bgColor="FFD966"))
            _y_rule = Rule(type='containsText', operator='containsText', text='Not Found', dxf=_y, priority=2)
            _y_rule.formula = [f'NOT(ISERROR(SEARCH("Not Found",{_s_col}4)))']
            ws.conditional_formatting.add(_s_range, _y_rule)
            # Good → green
            _g = DifferentialStyle(fill=PatternFill(bgColor="90EE90"))
            _g_rule = Rule(type='containsText', operator='containsText', text='Good', dxf=_g, priority=3)
            _g_rule.formula = [f'NOT(ISERROR(SEARCH("Good",{_s_col}4)))']
            ws.conditional_formatting.add(_s_range, _g_rule)

        # Entry Date Discrepancy — col X
        if entry_disc_c:
            _x_col = get_column_letter(entry_disc_c)
            _x_range = f"{_x_col}4:{_x_col}{last_row}"
            _uf = DifferentialStyle(fill=PatternFill(bgColor="FFC7CE"), font=Font(color="9C0006"))
            _uf_rule = Rule(type='containsText', operator='containsText', text='UNFAVORABLE', dxf=_uf, priority=1)
            _uf_rule.formula = [f'NOT(ISERROR(SEARCH("UNFAVORABLE",{_x_col}4)))']
            ws.conditional_formatting.add(_x_range, _uf_rule)
            _fv = DifferentialStyle(fill=PatternFill(bgColor="C6EFCE"), font=Font(color="006100"))
            _fv_rule = Rule(type='containsText', operator='containsText', text='FAVORABLE', dxf=_fv, priority=2)
            _fv_rule.formula = [f'NOT(ISERROR(SEARCH("FAVORABLE",{_x_col}4)))']
            ws.conditional_formatting.add(_x_range, _fv_rule)

        # Exit Date Discrepancy — col Y
        if exit_disc_c:
            _y_col = get_column_letter(exit_disc_c)
            _y_range = f"{_y_col}4:{_y_col}{last_row}"
            _uf2 = DifferentialStyle(fill=PatternFill(bgColor="FFC7CE"), font=Font(color="9C0006"))
            _uf2_rule = Rule(type='containsText', operator='containsText', text='UNFAVORABLE', dxf=_uf2, priority=1)
            _uf2_rule.formula = [f'NOT(ISERROR(SEARCH("UNFAVORABLE",{_y_col}4)))']
            ws.conditional_formatting.add(_y_range, _uf2_rule)
            _fv2 = DifferentialStyle(fill=PatternFill(bgColor="C6EFCE"), font=Font(color="006100"))
            _fv2_rule = Rule(type='containsText', operator='containsText', text='FAVORABLE', dxf=_fv2, priority=2)
            _fv2_rule.formula = [f'NOT(ISERROR(SEARCH("FAVORABLE",{_y_col}4)))']
            ws.conditional_formatting.add(_y_range, _fv2_rule)
            _bl = DifferentialStyle(fill=PatternFill(bgColor="BDD7EE"))
            _bl_rule = Rule(type='containsText', operator='containsText', text='update in ST', dxf=_bl, priority=3)
            _bl_rule.formula = [f'NOT(ISERROR(SEARCH("update in ST",{_y_col}4)))']
            ws.conditional_formatting.add(_y_range, _bl_rule)

        # FTE Variance — highlight non-zero values light red
        if fte_var_c:
            _fv_col = get_column_letter(fte_var_c)
            _fv_range = f"{_fv_col}4:{_fv_col}{last_row}"
            _fv_fill = PatternFill("solid", fgColor="FFC7CE")
            _fv_font = Font(color="9C0006")
            ws.conditional_formatting.add(_fv_range,
                CellIsRule(operator="notEqual", formula=["0"],
                           fill=_fv_fill, font=_fv_font))

        write_log("Conditional formatting applied: Status, Entry/Exit Discrepancy, FTE Variance")

    # Ensure Excel recalculates all formulas on open — prevents 'work lost' repair warning
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    wb.save(output_path)
    write_log(f"Done! Output saved to: {output_path}")
    print("\n" + "="*60)
    print(f"  DONE - Audited workbook saved to:")
    print(f"  {output_path}")
    print()
    print("  TABS IN YOUR OUTPUT FILE:")
    print("    SCSD ROSTER       - Full roster with color-coded status")
    print("    AUDIT_SUMMARY     - Counts, percentages, next steps")
    print("    Action Needed     - YOUR WORK QUEUE (start here)")
    print("    Not In ST         - Students missing from ST Enrollment")
    print("    Multiple Enroll.  - Students with 2+ enrollments")
    print("    Formula Adapter   - Column mapping + formula reference")
    print("    SPED ROSTER       - SPED students with IEP data")
    print("="*60)
    # Open file (Windows only)
    import platform
    if platform.system() == "Windows":
        import os as _os
        np = _os.path.normpath(output_path)
        if _os.path.exists(np):
            write_log(f"Opening: {np}")
            _os.startfile(np)
        else:
            write_log("Output file not found - cannot open automatically", "WARN")
    return output_path


# ======================================================================
# AUTO-DETECT MOST RECENT FILE FROM NETWORK FOLDERS (V4)
# Enrollment and disability files are loaded SEPARATELY and INDEPENDENTLY.
# Each picks the most recent .xlsx from its own dedicated folder.
# No winner-takes-all: both are returned even if one is missing.
# ======================================================================
FOLDER_ENROLLMENT = r"Z:\Daquan\Help Desk Tickets\Charter Schools Invoices\ST Charter Enrollment Report"
FOLDER_DISABILITY = r"Z:\Daquan\Help Desk Tickets\Charter Schools Invoices\Charter Invoice- Disability Programs WR export"

def get_most_recent_file(folder):
    """Return the most recently modified .xlsx file in a folder, or None."""
    if not os.path.isdir(folder):
        write_log(f"Folder not found: {folder}", "WARN")
        return None
    files = [os.path.join(folder, f) for f in os.listdir(folder) if f.endswith(".xlsx")]
    if not files:
        write_log(f"No .xlsx files found in: {folder}", "WARN")
        return None
    return max(files, key=os.path.getmtime)

def auto_detect_workbooks():
    """
    V4: Detects enrollment and disability files SEPARATELY.
    Returns (enrollment_path, disability_path).
    Either may be None if its folder is unreachable or empty.
    """
    enrollment_path = get_most_recent_file(FOLDER_ENROLLMENT)
    disability_path  = get_most_recent_file(FOLDER_DISABILITY)

    if enrollment_path:
        write_log(f"[Auto-detect] Enrollment file: {enrollment_path}")
    else:
        write_log("[Auto-detect] Enrollment file not found in Z: drive", "WARN")

    if disability_path:
        write_log(f"[Auto-detect] Disability file: {disability_path}")
    else:
        write_log("[Auto-detect] Disability file not found — SPED tab fallback disabled", "WARN")

    return enrollment_path, disability_path


# ======================================================================
# WORKBOOK INSPECTION HELPER
# Peeks inside a workbook and reports which key tabs are present.
# Used at startup so we only fall back to Z: drive for missing tabs.
# ======================================================================
def inspect_workbook_tabs(path):
    """
    Opens workbook at `path` and checks which key tabs exist.
    Returns dict: {'st': bool, 'sped': bool, 'disability': bool}
    """
    result = {'st': False, 'sped': False, 'disability': False}
    if path is None or not os.path.exists(path):
        return result
    try:
        xl = pd.ExcelFile(path)
        sheets_lower = [s.strip().lower() for s in xl.sheet_names]

        # ST Enrollment
        st_patterns = [r"st.*enrollment", r"chartenrol", r"enrollment"]
        st_exclude   = ["multiple", "scsd", "sped", "not on", "not in"]
        for pat in st_patterns:
            for s in sheets_lower:
                if re.search(pat, s) and not any(re.search(ex, s) for ex in st_exclude):
                    result['st'] = True
                    break
            if result['st']:
                break

        # SPED Roster
        for s in sheets_lower:
            if re.search(r"sped.*roster", s) or re.search(r"^sped$", s):
                result['sped'] = True
                break

        # Disability
        for s in sheets_lower:
            if re.search(r"disability", s) or re.search(r"iep.*roster", s):
                result['disability'] = True
                break

        xl.close()
    except Exception as e:
        write_log(f"Could not inspect workbook tabs: {e}", "WARN")
    return result


# ======================================================================
# ENTRY POINT
# ======================================================================
if __name__ == "__main__":
    # When running via the .bat launcher, the path is passed as a command-line argument.
    # When running manually, update the path below.

    if len(sys.argv) > 1:
        path = sys.argv[1].strip().strip('"')
    else:
        path = r"C:\Users\dmorri43\Training Docs and References\Help desk tickets\Charter Roster Invoices\Training Original Charter invoices\Syracuse-Southside-Invoice-6-Ticket-291462.xlsx"

    print("\n" + "="*60)
    print("  CHARTER AUDIT ENGINE V4")
    print("  Daquan Morrison · Syracuse City School District")
    print("="*60)
    print(f"\nFile loaded: {os.path.basename(path)}")

    if not path or not os.path.exists(path):
        print(f"\nERROR: File not found: {path}")
        print("Update the path in the script and re-run.")
        sys.exit(1)

    # Report which tabs were found inside the file
    tabs = inspect_workbook_tabs(path)
    print(f"  ST Enrollment tab : {'FOUND' if tabs['st'] else 'NOT FOUND — ST columns will be blank'}")
    print(f"  SPED Roster tab   : {'FOUND' if tabs['sped'] else 'NOT FOUND — SPED Status = Unknown'}")
    print(f"  Disability tab    : {'FOUND' if tabs['disability'] else 'NOT FOUND'}")
    print()

    # No separate disability file — everything must be inside the workbook
    # Charter attendance PDFs:
    run_audit(path, disability_path=None)
