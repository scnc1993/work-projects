"""
============================================================
CHARTER WORKSHEET DATA VALIDATION ENGINE  (Python / pandas)
SPED / SCSD Roster vs SchoolTool Audit Engine
------------------------------------------------------------
Layout-agnostic: auto sheet detection, auto header-row
detection, fuzzy column matching, flexible date parsing.
Produces the 13 calculated audit fields and writes a
multi-tab .xlsx report with severity highlighting.

Mirrors sql/charter_audit_engine.sql and the R version in
r/charter_audit_engine.R (identical field definitions).

Join model: rosters drive the audit (roster LEFT JOIN
SchoolTool on NYSSIS, both cast to text). StudentID columns
are validation inputs, NOT join keys.

Author: scnc1993 / Thrivora Holdings LLC

Requires: pandas, numpy, openpyxl
    pip install pandas numpy openpyxl
============================================================
"""

from __future__ import annotations

import re
from datetime import datetime, date

import numpy as np
import pandas as pd

ENGINE_VERSION = "Charter Worksheet Data Validation Engine (Python) v2026.06.06"
print(f"\n{ENGINE_VERSION} loaded")

# ============================================================
# 1. AUTO SHEET DETECTION
# ============================================================
_SHEET_PATTERNS = [
    r"^SPED ROSTER$", r"SPED ROSTER", r"SCSD SPED ROSTER", r"SPED",
    r"^SCSD ROSTER$", r"SCSD ROSTER", r"SCSD GEN ED", r"SCSD",
    r"^GEN ROSTER$", r"GEN ROSTER", r"GENERAL ED ROSTER", r"GEN ED ROSTER",
    r"ROSTER", r"ENROLL", r"STUDENT", r"INVOICE",
]


def detect_roster_sheet(path: str) -> str:
    xls = pd.ExcelFile(path)
    sheets = xls.sheet_names
    for pat in _SHEET_PATTERNS:
        for s in sheets:
            if re.search(pat, str(s), flags=re.IGNORECASE):
                return s
    return sheets[0]


# ============================================================
# 2. AUTO HEADER-ROW DETECTION
# ============================================================
def detect_header_row(path: str, sheet: str) -> int:
    raw = pd.read_excel(path, sheet_name=sheet, header=None, nrows=10, dtype=str)
    for i in range(len(raw)):
        if raw.iloc[i].notna().any():
            return i
    return 0


# ============================================================
# 3. FUZZY COLUMN DETECTOR
# ============================================================
def find_col(df: pd.DataFrame, patterns: list[str]) -> str | None:
    cols = list(df.columns)
    for pat in patterns:
        for c in cols:
            if re.search(pat, str(c), flags=re.IGNORECASE):
                return c
    return None


# ============================================================
# 4 & 5. NORMALIZERS
# ============================================================
_BLANKS = {"", "NA", "N/A", "NULL"}
_DOB_FORMATS = [
    "%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%m/%d/%y", "%m-%d-%y",
    "%d-%b-%Y", "%d-%b-%y", "%d/%m/%Y", "%Y/%m/%d", "%Y.%m.%d",
]


def normalize_name(x) -> str | None:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    s = re.sub(r"[^A-Z ]", "", str(x).upper())
    return re.sub(r"\s+", " ", s).strip()


def normalize_dob(x) -> str | None:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    s = str(x).strip().upper()
    if s in _BLANKS:
        return None
    # Excel serial
    try:
        num = float(s)
        if 20000 < num < 60000:
            d = date(1899, 12, 30) + pd.to_timedelta(int(num), unit="D").to_pytimedelta()
            return d.isoformat()
    except (ValueError, OverflowError):
        pass
    # Compact YYYYMMDD
    if re.fullmatch(r"\d{8}", s):
        try:
            return datetime.strptime(s, "%Y%m%d").date().isoformat()
        except ValueError:
            pass
    # SAS 14FEB2026
    if re.fullmatch(r"\d{1,2}[A-Z]{3}\d{4}", s):
        try:
            return datetime.strptime(s, "%d%b%Y").date().isoformat()
        except ValueError:
            pass
    for fmt in _DOB_FORMATS:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def normalize_value(x) -> str | None:
    """Trim, uppercase, collapse spaces, strip trailing .0, keep leading
    zeros, whitelist [^0-9A-Z/: -], parse dates / Excel serials."""
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    s = str(x).strip().upper()
    if s in _BLANKS:
        return None
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\.0+$", "", s)
    s = re.sub(r"[^0-9A-Z/: \-]", "", s)

    # timestamp -> date
    for fmt in ("%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    # date formats
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%m/%d/%y", "%m-%d-%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    # Excel serial / pure integer
    if re.fullmatch(r"[0-9]+", s):
        num = int(s)
        if 20000 < num < 60000:
            d = date(1899, 12, 30) + pd.to_timedelta(num, unit="D").to_pytimedelta()
            return d.isoformat()
        return str(num)
    return s


def _to_date(s):
    if s is None or s == "" or pd.isna(s):
        return pd.NaT
    return pd.to_datetime(s, errors="coerce")


def diff_days(roster_series: pd.Series, st_series: pd.Series) -> pd.Series:
    """roster - schooltool, in days; NaN if either missing."""
    rd = pd.to_datetime(roster_series, errors="coerce")
    sd = pd.to_datetime(st_series, errors="coerce")
    return (rd - sd).dt.days


# ============================================================
# 6. READ A ROSTER (auto sheet + header, all text)
# ============================================================
def read_roster(path: str, sheet: str | None = None) -> pd.DataFrame:
    if sheet is None:
        sheet = detect_roster_sheet(path)
    hdr = detect_header_row(path, sheet)
    return pd.read_excel(path, sheet_name=sheet, header=hdr, dtype=str)


# ============================================================
# 7. CORE AUDIT — produces the 13 calculated fields
# ============================================================
def charter_audit(roster_df: pd.DataFrame, st_df: pd.DataFrame,
                  school_label: str = "Roster",
                  tolerance_days: int = 5) -> pd.DataFrame:

    def col(df, patterns):
        c = find_col(df, patterns)
        if c is not None:
            return df[c].map(normalize_value)
        return pd.Series([None] * len(df), index=df.index)

    def raw_col(df, patterns):
        c = find_col(df, patterns)
        if c is not None:
            return df[c]
        return pd.Series([None] * len(df), index=df.index)

    # --- roster keys ---
    r = pd.DataFrame(index=roster_df.index)
    r["_r_ny"] = col(roster_df, [r"NYSSIS", r"NYS ?ID", r"STATE ?ID", r"UNIVERSAL"])
    r["_r_sid"] = col(roster_df, [r"STUDENT.?ID", r"LOCAL.?ID", r"STUDENT NUMBER"])
    r["_r_dis"] = raw_col(roster_df, [r"DISABILITY", r"CLASSIFICATION", r"IEP", r"SPED"])
    r["_r_ent"] = col(roster_df, [r"ENTRY", r"ENROLL", r"START"])
    r["_r_exit"] = col(roster_df, [r"EXIT", r"WITHDRAW", r"DISENROLL", r"END"])

    # --- SchoolTool keys ---
    s = pd.DataFrame(index=st_df.index)
    s["_st_ny"] = col(st_df, [r"NYSSIS", r"NYS ?ID", r"STATE ?ID", r"UNIVERSAL"])
    s["_st_sid"] = col(st_df, [r"STUDENT.?ID", r"LOCAL.?ID", r"STUDENT NUMBER"])
    s["_st_uid"] = col(st_df, [r"UNIVERSAL.?STUDENT", r"STUDENT_?UNIVERSAL"])
    s["_st_dis"] = raw_col(st_df, [r"DISABILITY", r"CLASSIFICATION", r"IEP", r"SPED"])
    s["_st_ent"] = col(st_df, [r"ENTRY", r"ENROLL", r"START"])
    s["_st_exit"] = col(st_df, [r"EXIT", r"WITHDRAW", r"DISENROLL", r"END"])
    # collapse SchoolTool to one row per NYSSIS (first match)
    s_keyed = s.dropna(subset=["_st_ny"]).drop_duplicates(subset=["_st_ny"])

    # roster-driven LEFT JOIN on NYSSIS (text)
    out = roster_df.copy().reset_index(drop=True)
    r = r.reset_index(drop=True)
    j = r.merge(s_keyed, left_on="_r_ny", right_on="_st_ny", how="left")

    def blank_to_na(series):
        return series.where(series.notna() & (series != ""), other=np.nan)

    st_sid = blank_to_na(j["_st_sid"])
    st_uid = blank_to_na(j["_st_uid"])
    r_sid = blank_to_na(j["_r_sid"])

    # 1. resolved_studentid — COALESCE waterfall
    resolved = st_sid.combine_first(st_uid).combine_first(r_sid).fillna("")
    out["resolved_studentid"] = resolved

    # 3. id_source
    out["id_source"] = np.select(
        [st_sid.notna(), st_uid.notna(), r_sid.notna()],
        ["SchoolTool StudentID", "Universal StudentID", "NYSSID Mapping"],
        default="No Match",
    )

    # 2. id_match_status
    matched = j["_st_ny"].notna()
    out["id_match_status"] = np.select(
        [~matched, resolved == ""],
        ["No Roster Match", "Missing Student ID Match"],
        default=f"Matched - {school_label}",
    )

    # 4. is_iep_flag (+ ST iep helper)
    def iep_flag(series):
        v = series.fillna("").astype(str).str.strip()
        return np.where(v == "", "Gen Ed", "IEP")

    out["is_iep_flag"] = iep_flag(j["_r_dis"])
    st_iep = iep_flag(j["_st_dis"])

    # 5. roster_vs_disability_check
    r_iep = out["is_iep_flag"].values
    out["roster_vs_disability_check"] = np.select(
        [
            ~matched.values,
            (r_iep == "IEP") & (st_iep == "IEP"),
            (r_iep == "Gen Ed") & (st_iep == "Gen Ed"),
            (r_iep == "IEP") & (st_iep == "Gen Ed"),
            (r_iep == "Gen Ed") & (st_iep == "IEP"),
        ],
        [
            "No SchoolTool Record",
            "Consistent SPED",
            "Consistent Gen Ed",
            "Mismatch - Roster SPED / ST Gen Ed",
            "Mismatch - Roster Gen Ed / ST SPED",
        ],
        default="Unknown",
    )

    # 6. entry_diff_days
    out["entry_diff_days"] = diff_days(j["_r_ent"], j["_st_ent"])
    ent_missing = j["_r_ent"].isna() | (j["_r_ent"] == "") | j["_st_ent"].isna() | (j["_st_ent"] == "")
    ed = out["entry_diff_days"]

    # 7. entry_status_category
    out["entry_status_category"] = np.select(
        [ent_missing.values, (ed == 0).values, (ed.abs() <= tolerance_days).values],
        ["Missing Date", "Exact Match", "Within Tolerance"],
        default="Finance Risk",
    )

    # 8. exit_diff_days
    out["exit_diff_days"] = diff_days(j["_r_exit"], j["_st_exit"])
    xd = out["exit_diff_days"]
    r_exit_missing = j["_r_exit"].isna() | (j["_r_exit"] == "")
    st_exit_missing = j["_st_exit"].isna() | (j["_st_exit"] == "")
    exit_missing = r_exit_missing | st_exit_missing

    # 9. exit_status_category
    out["exit_status_category"] = np.select(
        [exit_missing.values, (xd == 0).values, (xd.abs() <= tolerance_days).values],
        ["Missing Date", "Exact Match", "Within Tolerance"],
        default="Finance Risk",
    )

    # 10. exit_issue_type
    out["exit_issue_type"] = np.select(
        [
            (r_exit_missing & st_exit_missing).values,
            r_exit_missing.values,
            st_exit_missing.values,
            (xd > tolerance_days).values,
            (xd < -tolerance_days).values,
            (xd.abs() <= tolerance_days).values,
        ],
        [
            "Both Dates Missing",
            "Missing Roster Exit",
            "Missing ST Exit",
            "Roster Exit Late (>5)",
            "Roster Exit Early (<-5)",
            "Within Tolerance",
        ],
        default="No Issue",
    )

    # 11. final_validation_status
    entry_cat = out["entry_status_category"]
    exit_cat = out["exit_status_category"]
    rvd = out["roster_vs_disability_check"]
    out["final_validation_status"] = np.select(
        [
            (out["id_match_status"] == "No Roster Match") | (resolved == ""),
            (entry_cat == "Missing Date") | (exit_cat == "Missing Date"),
            (entry_cat == "Finance Risk") | (exit_cat == "Finance Risk"),
            rvd.isin(["Mismatch - Roster SPED / ST Gen Ed",
                      "Mismatch - Roster Gen Ed / ST SPED"]),
            (entry_cat == "Within Tolerance") | (exit_cat == "Within Tolerance"),
        ],
        [
            "Invalid - No ID Match",
            "Incomplete - Missing Dates",
            "Invalid - Date Mismatch",
            "Valid with Warning",
            "Valid with Warning",
        ],
        default="Valid",
    )

    # 12. severity
    fvs = out["final_validation_status"]
    out["severity"] = np.select(
        [
            fvs.isin(["Invalid - No ID Match", "Invalid - Date Mismatch"]),
            fvs == "Incomplete - Missing Dates",
            fvs == "Valid with Warning",
        ],
        ["Critical", "High", "Low"],
        default="None",
    )

    # 13. recommended_action
    eit = out["exit_issue_type"]
    out["recommended_action"] = np.select(
        [
            (out["id_match_status"] == "No Roster Match").values,
            (resolved == "").values,
            (eit == "Missing Roster Exit").values,
            (eit == "Missing ST Exit").values,
            (eit == "Both Dates Missing").values,
            (eit == "Roster Exit Late (>5)").values,
            (eit == "Roster Exit Early (<-5)").values,
            (entry_cat == "Finance Risk").values,
            (entry_cat == "Missing Date").values,
            (rvd == "Mismatch - Roster SPED / ST Gen Ed").values,
            (rvd == "Mismatch - Roster Gen Ed / ST SPED").values,
        ],
        [
            "Locate student in SchoolTool; confirm NYSSIS on roster.",
            "Resolve Student ID via NYSSIS mapping or Universal ID.",
            "Add exit/withdrawal date to roster.",
            "Enter exit date in SchoolTool.",
            "Verify enrollment status; both exit dates absent.",
            "Reconcile roster exit date - billed beyond ST exit.",
            "Reconcile roster exit date - earlier than ST exit.",
            "Reconcile entry dates - FTE/billing risk.",
            "Add missing entry/enrollment date.",
            "Confirm IEP classification; roster says SPED.",
            "Confirm IEP classification; ST says SPED.",
        ],
        default="No action - record valid.",
    )

    out["school"] = school_label
    return out


# ============================================================
# 8. EXPORT — multi-tab workbook with severity highlighting
# ============================================================
def export_audit(audit_df: pd.DataFrame,
                 out_file: str = "CHARTER_AUDIT_REPORT.xlsx") -> None:
    from openpyxl.styles import PatternFill

    fills = {
        "Critical": PatternFill("solid", fgColor="F4CCCC"),
        "High": PatternFill("solid", fgColor="FCE5CD"),
        "Low": PatternFill("solid", fgColor="FFF2CC"),
    }

    exceptions = audit_df[audit_df["severity"].isin(["Critical", "High", "Low"])]
    summary = (
        audit_df.groupby(["severity", "final_validation_status"])
        .size().reset_index(name="count")
        .sort_values("count", ascending=False)
    )

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        audit_df.to_excel(writer, sheet_name="Full_Audit", index=False)
        exceptions.to_excel(writer, sheet_name="Exceptions", index=False)
        summary.to_excel(writer, sheet_name="Severity_Summary", index=False)

        ws = writer.sheets["Full_Audit"]
        ws.freeze_panes = "A2"
        sev_idx = list(audit_df.columns).index("severity")
        for row_offset, sev in enumerate(audit_df["severity"].values, start=2):
            fill = fills.get(sev)
            if fill is not None:
                for col_idx in range(1, len(audit_df.columns) + 1):
                    ws.cell(row=row_offset, column=col_idx).fill = fill

    print(f"Audit report saved to: {out_file}")


# ============================================================
# 9. MAIN WRAPPER
# ============================================================
def run_charter_audit(roster_file: str,
                      schooltool_file: str,
                      school_label: str = "Roster",
                      out_file: str = "CHARTER_AUDIT_REPORT.xlsx") -> pd.DataFrame:
    roster_df = read_roster(roster_file)
    st_df = read_roster(schooltool_file)

    audit_df = charter_audit(roster_df, st_df, school_label=school_label)
    export_audit(audit_df, out_file)

    print("\n===== CHARTER AUDIT COMPLETE =====")
    print(f"Rows audited:  {len(audit_df)}")
    print(f"Critical:      {(audit_df['severity'] == 'Critical').sum()}")
    print(f"High:          {(audit_df['severity'] == 'High').sum()}")
    print(f"Low:           {(audit_df['severity'] == 'Low').sum()}")
    print(f"Valid (None):  {(audit_df['severity'] == 'None').sum()}")
    return audit_df


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Charter SPED/Roster audit engine")
    parser.add_argument("roster", help="Path to roster .xlsx")
    parser.add_argument("schooltool", help="Path to SchoolTool export .xlsx")
    parser.add_argument("--school", default="Roster", help="School label")
    parser.add_argument("--out", default="CHARTER_AUDIT_REPORT.xlsx", help="Output file")
    args = parser.parse_args()

    run_charter_audit(args.roster, args.schooltool,
                      school_label=args.school, out_file=args.out)
