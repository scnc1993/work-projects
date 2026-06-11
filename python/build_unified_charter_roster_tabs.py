"""
rebuild_with_st_tab.py
Daquan Morrison · Data Analyst · Syracuse City School District

Fixes:
  1. NYSSIS=0/blank (4 SS rows) → patch from ST by name+DOB
  2. Blank Last Name (4 SS rows where last = "0") → patch from ST
  3. Blank Student Status (249 OnTech + 3 SS) → classify from ST dates
  4. Blank School Name (3 SAS + 1 SASCCS) → patch from ST
  5. Blank Student ID (5+ SASCCS) → patch from ST by NYSSIS
  6. Add TAB 6: ST Enroll Export (combined all 4 charters, deduplicated)
  7. Formula Version tab uses VLOOKUP against 'ST Enroll Export'!A:K
"""
import openpyxl, zipfile, os
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.differential import DifferentialStyle
from datetime import datetime, date
from collections import defaultdict
import xml.etree.ElementTree as ET

BASE = "/home/user/workspace/uploaded_attachments/8dc6b308dcc94470ad3ac69527bb318b/"
OUT  = "/home/user/workspace/all_charters_unified_roster_25-26_REBUILT.xlsx"

# ── helpers ──────────────────────────────────────────────────────────────────
def sv(v):
    if v is None: return ''
    return str(v).replace('.0','').strip()

def dv(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    return None

DATE_FMT = 'mm/dd/yyyy'
DATE_COLS = {2, 9, 11, 12}   # Charter Month, DOB, Enroll Date, Withdraw Date

# ── load one ST export ────────────────────────────────────────────────────────
def load_st(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['ST Enroll Export']
    h  = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    rows = []
    for r in range(2, ws.max_row+1):
        row = {h[i]: ws.cell(r, i+1).value for i in range(len(h))}
        if any(row.get(k) for k in ['Student_UniversalStudentID','StudentID','FirstName']):
            rows.append(row)
    return rows

print("Loading ST Enroll from all 4 charters...")
all_st_raw = []
for f in ["southside_291462_unified-roster.xlsx",
          "sas_291463_unified-roster.xlsx",
          "sasccs_291464_unified-roster.xlsx",
          "ontech_final_unified-roster.xlsx"]:
    rows = load_st(BASE + f)
    all_st_raw.extend(rows)
    print(f"  {f}: {len(rows)} ST rows")

# deduplicate on NYSSIS+SID+StartDate+School
seen_keys = set()
combined_st = []
for row in all_st_raw:
    nyssis = sv(row.get('Student_UniversalStudentID'))
    sid    = sv(row.get('StudentID'))
    start  = str(row.get('StudentEnrollment_StartDate') or '')[:10]
    school = sv(row.get('SchoolName'))
    key    = f"{nyssis}|{sid}|{start}|{school}"
    if key not in seen_keys:
        seen_keys.add(key)
        combined_st.append(row)

print(f"Combined unique ST rows: {len(combined_st)}")

# ── ST lookup dicts ───────────────────────────────────────────────────────────
st_by_nyssis = defaultdict(list)
st_by_sid    = defaultdict(list)
st_by_name   = defaultdict(list)

for row in combined_st:
    nyssis = sv(row.get('Student_UniversalStudentID'))
    sid    = sv(row.get('StudentID'))
    fname  = sv(row.get('FirstName')).upper()
    lname  = sv(row.get('LastName')).upper()
    if nyssis and nyssis not in ('0',''):
        st_by_nyssis[nyssis].append(row)
    if sid and sid not in ('0',''):
        st_by_sid[sid].append(row)
    if fname and lname:
        st_by_name[(lname, fname)].append(row)

# ── classify logic ────────────────────────────────────────────────────────────
CHARTER_KW = {
    'Southside 291462': ['southside academy'],
    'SAS 291463':       ['syracuse academy of science','sas'],
    'SASCCS 291464':    ['citizenship & science','citizenship and science','sasccs'],
    'OnTech 291465':    ['ontech','on tech'],
}

def school_ok(school_str, charter):
    sl = school_str.lower()
    return any(kw in sl for kw in CHARTER_KW.get(charter, []))

def classify(enroll_dt, withdraw_dt, st_rows, charter):
    if not st_rows:
        return 'Not in ST System', \
               'Not pulling up in ST. Need more student information.'
    charter_rows = [r for r in st_rows if school_ok(sv(r.get('SchoolName','')), charter)]
    if not charter_rows:
        other = sv(st_rows[0].get('SchoolName',''))
        o_s   = dv(st_rows[0].get('StudentEnrollment_StartDate'))
        o_e   = dv(st_rows[0].get('StudentEnrollment_EndDate'))
        ds    = o_s.strftime('%-m/%-d/%y') if o_s else ''
        ds   += f" - {o_e.strftime('%-m/%-d/%y')}" if o_e else ('-current' if ds else '')
        return 'No Good', \
               f"ST states: student started {other} {ds}.  Need charter attendance record to change in ST."
    charter_rows.sort(
        key=lambda r: dv(r.get('StudentEnrollment_StartDate')) or date(2000,1,1),
        reverse=True)
    best      = charter_rows[0]
    st_start  = dv(best.get('StudentEnrollment_StartDate'))
    st_end    = dv(best.get('StudentEnrollment_EndDate'))
    st_school = sv(best.get('SchoolName',''))
    notes, no_good = [], False
    if enroll_dt and st_start and abs((enroll_dt - st_start).days) > 3:
        no_good = True
        diff = abs((enroll_dt - st_start).days)
        notes.append(
            f"ST states: student enrolled at {st_school} "
            f"{st_start.strftime('%-m/%-d/%y')} . "
            f"Entry date is {diff}> discrepancy. "
            f"Need charter attendance record to change in ST.")
    if not st_start and enroll_dt:
        no_good = True
        notes.append("No enrollment record in ST. Need charter attendance record to change in ST.")
    if withdraw_dt and not st_end:
        notes.append(
            f"Charter has exit date {withdraw_dt.strftime('%-m/%-d/%y')} "
            f"but no ST exit — this will update in ST.")
    if withdraw_dt and st_end and abs((withdraw_dt - st_end).days) > 3:
        no_good = True
        notes.append(
            f"Charter exit {withdraw_dt.strftime('%-m/%-d/%y')} does not match "
            f"ST exit {st_end.strftime('%-m/%-d/%y')}. This will update in ST.")
    return ('No Good' if no_good else 'Good'), (' '.join(notes) if notes else None)

# ── load existing output ──────────────────────────────────────────────────────
print("Loading existing output file...")
wb_src = openpyxl.load_workbook(OUT, data_only=True)
ws_src = wb_src['All Charters Unified']
headers = [ws_src.cell(1,c).value for c in range(1, ws_src.max_column+1)]
col = {h: i for i,h in enumerate(headers) if h}  # 0-based

I_CHARTER  = col['Charter_Source Ticket #']
I_NYSSIS   = col['NYSSIS ID']
I_SID      = col['Student ID']
I_LNAME    = col['Last Name']
I_FNAME    = col['First Name']
I_SCHOOL   = col['School Name']
I_STATUS   = col['Student Status']
I_NOTES    = col['Notes for Patrick']
I_ENROLL   = col['Enroll Date']
I_WITHDRAW = col['Withdraw Date']
I_SPED     = col['SPED_Flag']

# read all data rows
data = []
for r in range(2, ws_src.max_row+1):
    row = [ws_src.cell(r,c).value for c in range(1, ws_src.max_column+1)]
    if not any(row): continue
    data.append(row)

print(f"  Loaded {len(data)} data rows")

# ── apply patches ─────────────────────────────────────────────────────────────
print("Patching data...")
fixes = defaultdict(int)

for row in data:
    charter  = sv(row[I_CHARTER])
    nyssis   = sv(row[I_NYSSIS])
    sid      = sv(row[I_SID])
    lname    = sv(row[I_LNAME])
    fname    = sv(row[I_FNAME])
    school   = sv(row[I_SCHOOL])
    status   = sv(row[I_STATUS])
    enroll   = dv(row[I_ENROLL])
    withdraw = dv(row[I_WITHDRAW])

    # find ST match
    st_rows = []
    if nyssis and nyssis not in ('0','','Not Found'):
        st_rows = st_by_nyssis.get(nyssis, [])
    if not st_rows and sid and sid not in ('0','','Not Found'):
        st_rows = st_by_sid.get(sid, [])
    if not st_rows and fname and lname:
        st_rows = st_by_name.get((lname.upper(), fname.upper()), [])

    # 1. Fix NYSSIS = 0 / blank / Not Found
    if nyssis in ('0','','Not Found') and st_rows:
        found = sv(st_rows[0].get('Student_UniversalStudentID',''))
        if found and found not in ('0',''):
            row[I_NYSSIS] = found
            nyssis = found
            fixes['nyssis'] += 1

    # 2. Fix blank Last Name (stored as "0" in source)
    if lname in ('0','','None') and st_rows:
        found = sv(st_rows[0].get('LastName',''))
        if found:
            row[I_LNAME] = found
            fixes['lname'] += 1

    # 3. Fix blank School Name
    if not school or school in ('0','','None'):
        if st_rows:
            crows = [r for r in st_rows if school_ok(sv(r.get('SchoolName','')), charter)]
            if crows:
                found = sv(crows[0].get('SchoolName',''))
                if found:
                    row[I_SCHOOL] = found
                    school = found
                    fixes['school'] += 1

    # 4. Fix blank Student ID (mainly SASCCS)
    if not sid or sid in ('0','','Not Found','None'):
        if st_rows:
            found = sv(st_rows[0].get('StudentID',''))
            if found and found not in ('0',''):
                row[I_SID] = found
                sid = found
                fixes['sid'] += 1

    # 5. Fix blank Student Status (OnTech 249 rows + 3 SS rows)
    if not status or status in ('','None','NA'):
        new_status, new_notes = classify(enroll, withdraw, st_rows, charter)
        row[I_STATUS] = new_status
        if new_notes and sv(row[I_NOTES]) in ('','None','NA'):
            row[I_NOTES] = new_notes
        fixes['status'] += 1

print(f"  Patches: {dict(fixes)}")

# ── count issues remaining for verification ───────────────────────────────────
remaining = defaultdict(int)
from collections import Counter
status_counts = defaultdict(Counter)
for row in data:
    charter = sv(row[I_CHARTER])
    nyssis  = sv(row[I_NYSSIS])
    status  = sv(row[I_STATUS])
    school  = sv(row[I_SCHOOL])
    sid     = sv(row[I_SID])
    lname   = sv(row[I_LNAME])
    status_counts[charter][status] += 1
    if not nyssis or nyssis in ('0','Not Found','None'): remaining['nyssis_bad'] += 1
    if not status or status in ('','None','NA'):          remaining['status_blank'] += 1
    if not school or school in ('0','','None'):           remaining['school_blank'] += 1
    if not sid or sid in ('0','Not Found','None',''):     remaining['sid_bad'] += 1
    if not lname or lname in ('0','','None'):             remaining['lname_bad'] += 1

print(f"\nRemaining issues after patch: {dict(remaining)}")
print("\nStatus breakdown by charter:")
for charter in sorted(status_counts):
    print(f"  {charter}:")
    for st, cnt in sorted(status_counts[charter].items()):
        print(f"    {cnt:4d}  {st}")

# ── styles ────────────────────────────────────────────────────────────────────
DARK   = PatternFill('solid', fgColor='1F3864')
TEAL   = PatternFill('solid', fgColor='17375E')
ORANGE = PatternFill('solid', fgColor='C55A11')
PURPLE = PatternFill('solid', fgColor='4B2D83')
GREEN  = PatternFill('solid', fgColor='375623')
NAVY   = PatternFill('solid', fgColor='203864')

GOOD_FILL   = PatternFill('solid', fgColor='C6EFCE')
NOGOOD_FILL = PatternFill('solid', fgColor='FFC7CE')
SPED_FILL   = PatternFill('solid', fgColor='FFF2CC')
GENED_FILL  = PatternFill('solid', fgColor='E2EFDA')
FLAG_FILL   = PatternFill('solid', fgColor='FFEB9C')

WHITE_BOLD  = Font(bold=True, color='FFFFFF')
GOOD_FONT   = Font(color='276221')
NOGOOD_FONT = Font(color='9C0006')
FLAG_FONT   = Font(color='9C5700', bold=True)

COL_W = [18,13,34,13,12,14,12,10,12,32,12,12,9,16,52,12]

def hdr_cell(ws, r, c, val, fill):
    cell = ws.cell(r, c, val)
    cell.fill = fill
    cell.font = WHITE_BOLD
    cell.alignment = Alignment(horizontal='left')
    return cell

def color_row(ws, r_idx, row):
    """Apply status and SPED colors to a data row (r_idx is 1-based sheet row)."""
    status = sv(row[I_STATUS])
    sped   = sv(row[I_SPED])
    # Status col (col 14 = index 13 = I_STATUS, 1-based = 14)
    sc = I_STATUS + 1
    if 'No Good' in status:
        ws.cell(r_idx, sc).fill = NOGOOD_FILL
        ws.cell(r_idx, sc).font = NOGOOD_FONT
    elif status == 'Good':
        ws.cell(r_idx, sc).fill = GOOD_FILL
        ws.cell(r_idx, sc).font = GOOD_FONT
    # SPED col (col 16)
    pc = I_SPED + 1
    if 'SPED' in sped:    ws.cell(r_idx, pc).fill = SPED_FILL
    elif 'Gen Ed' in sped: ws.cell(r_idx, pc).fill = GENED_FILL

def write_data_row(ws, r_idx, row):
    for c_idx, val in enumerate(row, 1):
        cell = ws.cell(r_idx, c_idx, val)
        if c_idx in DATE_COLS and val is not None:
            cell.number_format = DATE_FMT
    color_row(ws, r_idx, row)

def set_col_widths(ws, widths):
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

n = len(data)

# ── build workbook ────────────────────────────────────────────────────────────
print("\nBuilding workbook...")
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — All Charters Unified
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("All Charters Unified")
for c, h in enumerate(headers, 1):
    hdr_cell(ws1, 1, c, h, DARK)
for r_idx, row in enumerate(data, 2):
    write_data_row(ws1, r_idx, row)
set_col_widths(ws1, COL_W)
ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n+1}"
print(f"  Tab 1: {n} rows")

# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — Formula Version  (VLOOKUP against ST Enroll Export)
# ════════════════════════════════════════════════════════════════════════════
# ST Enroll Export column layout:
#   A=NYSSIS_ID  B=Student_ID  C=Last_Name  D=First_Name  E=DOB  F=Grade
#   G=School_Name  H=Enroll_Start  I=Enroll_End  J=Entry_Note  K=Exit_Note
# VLOOKUP(lookup_value, 'ST Enroll Export'!$A:$K, col_index, 0)
#   col index 1=NYSSIS, 2=SID, 3=Last, 4=First, 5=DOB, 7=School

ws2 = wb.create_sheet("Formula Version")
fv_hdrs = list(headers) + ['ST_NYSSIS_Lookup', 'ST_School_Lookup',
                            'ST_SID_Lookup', 'Match_Source']
for c, h in enumerate(fv_hdrs, 1):
    hdr_cell(ws2, 1, c, h, TEAL)

ST = "'ST Enroll Export'"

for r_idx, row in enumerate(data, 2):
    write_data_row(ws2, r_idx, row)
    # Col letters in THIS sheet (Formula Version):
    # D = NYSSIS ID (col 4), E = Student ID (col 5)
    D = f"D{r_idx}"   # NYSSIS ID
    E = f"E{r_idx}"   # Student ID

    # Col 17: ST_NYSSIS_Lookup
    # VLOOKUP Student ID in ST col B → return NYSSIS (col A, index 1 of B:K)
    # Fallback: VLOOKUP NYSSIS in ST col A → return NYSSIS (confirms it exists)
    f17 = (f"=IFERROR(VLOOKUP({E},{ST}!$B:$A,1,0),"
           f"IFERROR(VLOOKUP({D},{ST}!$A:$A,1,0),\"Not Found\"))")
    ws2.cell(r_idx, 17, f17)

    # Col 18: ST_School_Lookup — look up NYSSIS in ST col A, return School (col 7 of A:K)
    f18 = (f"=IFERROR(VLOOKUP({D},{ST}!$A:$K,7,0),"
           f"IFERROR(VLOOKUP({E},{ST}!$B:$K,6,0),\"Not Found\"))")
    ws2.cell(r_idx, 18, f18)

    # Col 19: ST_SID_Lookup — look up NYSSIS in ST col A, return SID (col 2 of A:K)
    f19 = (f"=IFERROR(VLOOKUP({D},{ST}!$A:$K,2,0),"
           f"IFERROR(VLOOKUP({E},{ST}!$B:$B,1,0),\"Not Found\"))")
    ws2.cell(r_idx, 19, f19)

    # Col 20: Match_Source
    f20 = (f"=IF(Q{r_idx}=\"Not Found\",\"No ST Match\","
           f"IF(Q{r_idx}={D},\"NYSSIS Match\",\"SID Match\"))")
    ws2.cell(r_idx, 20, f20)

set_col_widths(ws2, COL_W + [15,20,12,12])
ws2.freeze_panes = 'A2'
print(f"  Tab 2: Formula Version written")

# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — Copy & Paste
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Copy & Paste")
banner = ws3.cell(1, 1,
    f"COPY & PASTE READY  |  Select A3:P{n+2}, copy, paste VALUES ONLY into FY 25-26 master.")
banner.fill = PatternFill('solid', fgColor='FCE4D6')
banner.font = Font(bold=True, color='843C0C')
banner.alignment = Alignment(horizontal='left')
ws3.merge_cells(f'A1:{get_column_letter(len(headers))}1')
for c, h in enumerate(headers, 1):
    hdr_cell(ws3, 2, c, h, ORANGE)
for r_idx, row in enumerate(data, 3):
    write_data_row(ws3, r_idx, row)
set_col_widths(ws3, COL_W)
ws3.freeze_panes = 'A3'
print(f"  Tab 3: Copy & Paste written")

# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — Not Found in ST
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Not Found in ST")
nf_rows = []
for row in data:
    nyssis = sv(row[I_NYSSIS])
    status = sv(row[I_STATUS])
    if not nyssis or nyssis in ('0','Not Found','None') or status == 'Not in ST System':
        action = ('Look up NYSSIS in SchoolTool — enter in Update Log tab'
                  if not nyssis or nyssis in ('0','Not Found','None')
                  else 'Student not found in ST enrollment — verify with Patrick')
        nf_rows.append(list(row) + [action])

nf_banner = ws4.cell(1, 1,
    f"NOT FOUND IN ST  |  {len(nf_rows)} students unresolvable — look up manually.")
nf_banner.fill = PatternFill('solid', fgColor='FFC7CE')
nf_banner.font = Font(bold=True, color='9C0006')
ws4.merge_cells(f'A1:{get_column_letter(len(headers)+1)}1')
for c, h in enumerate(list(headers)+['Action Needed'], 1):
    hdr_cell(ws4, 2, c, h, PURPLE)
for r_idx, row in enumerate(nf_rows, 3):
    for c_idx, val in enumerate(row, 1):
        cell = ws4.cell(r_idx, c_idx, val)
        if c_idx in DATE_COLS and val is not None:
            cell.number_format = DATE_FMT
        cell.fill = PatternFill('solid', fgColor='FFE4E1')
set_col_widths(ws4, COL_W + [44])
ws4.freeze_panes = 'A3'
print(f"  Tab 4: Not Found in ST — {len(nf_rows)} rows")

# ════════════════════════════════════════════════════════════════════════════
# TAB 5 — Update Log
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Update Log")
ul_hdrs = ['NYSSIS ID','Student ID','Last Name','First Name','DOB',
           'Grade Level','School Name','Charter_Source Ticket #',
           'Resolved_By','Resolved_Date','Notes']
ul_banner = ws5.cell(1, 1,
    "UPDATE LOG  |  Paste resolved NYSSIS here. Col A=NYSSIS | Col B=Student ID. "
    "Formula Version auto-pulls resolved NYSSIS for matching Student ID.")
ul_banner.fill = PatternFill('solid', fgColor='E2EFDA')
ul_banner.font = Font(bold=True, color='375623')
ws5.merge_cells(f'A1:{get_column_letter(len(ul_hdrs))}1')
for c, h in enumerate(ul_hdrs, 1):
    hdr_cell(ws5, 2, c, h, GREEN)
ws5.freeze_panes = 'A3'
print(f"  Tab 5: Update Log written")

# ════════════════════════════════════════════════════════════════════════════
# TAB 6 — ST Enroll Export (combined, deduplicated, VLOOKUP-ready)
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("ST Enroll Export")
# Column layout (VLOOKUP-friendly — NYSSIS in A, SID in B):
# A  B         C          D           E    F      G            H             I           J            K
# NYSSIS_ID  Student_ID  Last_Name  First_Name  DOB  Grade  School_Name  Enroll_Start  Enroll_End  Entry_Note  Exit_Note
st_out_hdrs = ['NYSSIS_ID','Student_ID','Last_Name','First_Name','DOB','Grade',
               'School_Name','Enroll_Start','Enroll_End','Entry_Note','Exit_Note']
st_src_keys  = ['Student_UniversalStudentID','StudentID','LastName','FirstName',
                'Person_DOB','Grade','SchoolName',
                'StudentEnrollment_StartDate','StudentEnrollment_EndDate',
                'StudentEnrollment_EntryNote','StudentEnrollment_ExitNote']

st_banner = ws6.cell(1, 1,
    "ST ENROLL EXPORT — Combined SchoolTool data all 4 charters | "
    "VLOOKUP on Col A (NYSSIS_ID) or Col B (Student_ID) | "
    "Daquan Morrison · Data Analyst · Syracuse City School District")
st_banner.fill = NAVY
st_banner.font = WHITE_BOLD
ws6.merge_cells(f'A1:{get_column_letter(len(st_out_hdrs))}1')

for c, h in enumerate(st_out_hdrs, 1):
    hdr_cell(ws6, 2, c, h, NAVY)

ST_DATE_COLS = {5, 8, 9}
for r_idx, st_row in enumerate(combined_st, 3):
    for c_idx, key in enumerate(st_src_keys, 1):
        val = st_row.get(key)
        if key in ('Student_UniversalStudentID','StudentID'):
            v = str(val or '').replace('.0','').strip()
            val = v if v not in ('None','0','') else None
        cell = ws6.cell(r_idx, c_idx, val)
        if c_idx in ST_DATE_COLS and val is not None:
            cell.number_format = DATE_FMT

# Freeze on row 3 (row 1=banner, row 2=header, row 3+ = data)
ws6.freeze_panes = 'A3'
ws6.auto_filter.ref = f"A2:{get_column_letter(len(st_out_hdrs))}{len(combined_st)+2}"
ST_COL_W = [14,12,18,14,12,7,40,13,13,45,45]
set_col_widths(ws6, ST_COL_W)
print(f"  Tab 6: ST Enroll Export — {len(combined_st)} rows")

# ── save + strip dangling drawing refs ───────────────────────────────────────
print("\nSaving...")
tmp = OUT.replace('.xlsx','_tmp2.xlsx')
wb.save(tmp)

with zipfile.ZipFile(tmp,'r') as zin, zipfile.ZipFile(OUT,'w',zipfile.ZIP_DEFLATED) as zout:
    for name in zin.namelist():
        data_bytes = zin.read(name)
        if name.endswith('.rels') and b'drawing' in data_bytes.lower():
            try:
                root = ET.fromstring(data_bytes)
                to_rm = [el for el in root
                         if 'drawing' in el.get('Type','').lower()
                         or 'drawing' in el.get('Target','').lower()]
                for el in to_rm:
                    root.remove(el)
                data_bytes = (b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                              + ET.tostring(root, encoding='unicode').encode('utf-8'))
            except: pass
        zout.writestr(name, data_bytes)

os.remove(tmp)
print(f"Saved: {OUT}")

# ── final verification ────────────────────────────────────────────────────────
print("\n=== FINAL VERIFICATION ===")
wb_v = openpyxl.load_workbook(OUT, data_only=True)
ws_v = wb_v['All Charters Unified']
h_v  = [ws_v.cell(1,c).value for c in range(1, ws_v.max_column+1)]
cv   = {h: i+1 for i,h in enumerate(h_v) if h}

issues = defaultdict(int)
sc_v   = defaultdict(Counter)
for r in range(2, ws_v.max_row+1):
    charter = sv(ws_v.cell(r, cv['Charter_Source Ticket #']).value)
    if not charter: continue
    nyssis  = sv(ws_v.cell(r, cv['NYSSIS ID']).value)
    sid     = sv(ws_v.cell(r, cv['Student ID']).value)
    school  = sv(ws_v.cell(r, cv['School Name']).value)
    status  = sv(ws_v.cell(r, cv['Student Status']).value)
    lname   = sv(ws_v.cell(r, cv['Last Name']).value)
    sc_v[charter][status] += 1
    if not nyssis or nyssis in ('0','Not Found','None'): issues['nyssis_bad']  += 1
    if not status or status in ('','None','NA'):          issues['status_blank'] += 1
    if not school or school in ('0','','None'):           issues['school_blank'] += 1
    if not sid    or sid    in ('0','Not Found','None'):  issues['sid_bad']      += 1
    if not lname  or lname  in ('0','','None'):           issues['lname_bad']    += 1

print(f"Sheets: {wb_v.sheetnames}")
print(f"Remaining issues: {dict(issues)}")
print("\nStudent Status by charter:")
for charter in sorted(sc_v):
    print(f"  {charter}:")
    for st, cnt in sorted(sc_v[charter].items()):
        print(f"    {cnt:4d}  {st}")

# Count totals
total = sum(sum(v.values()) for v in sc_v.values())
print(f"\nTotal rows: {total}")
